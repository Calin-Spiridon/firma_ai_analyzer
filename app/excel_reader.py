from openpyxl import load_workbook


def extract_all_years(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    years = {}

    for col_index in range(1, len(header)):
        year = header[col_index]
        if isinstance(year, int):
            years[year] = {}

    for row in rows[1:]:
        indicator = row[0]

        for col_index in range(1, len(header)):
            year = header[col_index]

            if isinstance(year, int):
                value = row[col_index]
                years[year][indicator] = value

    return years